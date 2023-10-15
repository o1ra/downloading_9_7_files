import zipfile
import os.path
import xlrd
from PyPDF2 import PdfReader
from openpyxl.reader.excel import load_workbook

current_dir = os.path.dirname(os.path.abspath(__file__))
resources_files = os.path.join(current_dir, "resources")
file_dir = os.listdir(resources_files)
tmp = os.path.join(current_dir, "tmp")
zip_file = os.path.join(tmp, "test.zip")


def test_xls_from_zip():
    the_zip_file = zipfile.ZipFile(zip_file)
    res = the_zip_file.testzip()
    assert res is None


def test_pdf():
    with zipfile.ZipFile('tmp/test.zip', 'r') as myzip:
        with myzip.open('Python Testing with Pytest (Brian Okken).pdf') as pdf_t:
            pdf = PdfReader(pdf_t)
            page = pdf.pages[1]
            text = page.extract_text()
            assert 'Python Testing' in text

            number_of_page = len(pdf.pages)
            assert number_of_page == 256

            size_pdf = myzip.read('Python Testing with Pytest (Brian Okken).pdf')
            pdf_file_size = len(size_pdf)
            expected_file_size = 3035139
            assert pdf_file_size == expected_file_size


def test_txt():
    with zipfile.ZipFile('tmp/test.zip', 'r') as myzip:
        with myzip.open('Hello.txt') as txt_t:
            assert txt_t.read().decode('utf-8') == 'Hello world\n'


def test_xls():
    with zipfile.ZipFile('tmp/test.zip', 'r') as myzip:
        with myzip.open('file_example_XLS_10.xls') as xls_t:
            xls_content = xls_t.read()
            xls_workbook = xlrd.open_workbook(file_contents=xls_content)
            sheet = xls_workbook.sheet_by_index(0)
            expected_value = 1562
            actual_value = sheet.cell_value(1, 7)
            assert actual_value == expected_value


def test_xlsx():
    with zipfile.ZipFile('tmp/test.zip', 'r') as myzip:
        with myzip.open('file_example_XLSX_50.xlsx') as xlsx_t:
            xlsx_file = load_workbook(xlsx_t)
            sheet = xlsx_file.active
            value = sheet.cell(row=3, column=3).value
            assert value == 'Hashimoto'

